#!/usr/bin/env python 
# -*- coding: utf-8 -*- 
# @Time : 2020/5/20 14:03
# @Author : Lemon_Yun
# @QQ: 316093170
# Copyright：中华人民共和国

'''


'''

import openpyxl   #导入EXCEL第三方库
import requests  #导入 http第三方库
session = requests.session()   #将requests库的session模块赋值给变量，后期需要调用cookie
#读取测试用例数据的函数：
def read_data(filename,sheetname):
    wb = openpyxl.load_workbook(filename)
    sheet = wb[sheetname]
    max_row = sheet.max_row   #获取表格最大行数
    cases = []  #创建一个空列表
    for i in range(2,max_row+1):
        case = dict(
        case_id = sheet.cell(row=i, column=1).value,  # 获取编号
        url = sheet.cell(row=i,column=5).value,  #获取url
        data = sheet.cell(row=i,column=6).value,  #获取data
        expected_result = sheet.cell(row=i,column=7).value  #获取期望结果
        )  #一个用例存放到一个字典
        cases.append(case)   #把字典追加到上面创建的空列表中保存起来
    return cases #定义返回值
#写入测试结论的函数：
def wite_result(filename,sheetname,row,column,real_result):
    wb = openpyxl.load_workbook(filename)
    sheet = wb[sheetname]
    sheet.cell(row=row, column=column).value = real_result   #写入
    wb.save(filename)    #保存
#发送接口请求的函数：
def post_func(qcd_url,qcd_data):
    res = session.post(url = qcd_url,data = qcd_data)  #post方法发送接口请求，session为上面的变量，自动带cookies值
    result = res.json() #变量以字典数据格式保存，并赋值给result
    return result   #返回响应消息


#测试脚本开始：
def execute_func(filename,sheetname):      #封装下面已经写好的函数
    test_cases = read_data(filename,sheetname)   #调用读取测试用例的函数并赋值给变量
    for test_case in test_cases:        #将读取的数据分成一条条测试用例
        case_id = test_case.get('case_id')   #获取每一条测试用例的case_id
        url = test_case.get('url')        #获取每一条测试用例的url
        data = test_case.get('data')     #获取每一条测试用例的请求参数  从excel中读取的数据都是文本文档数据格式
        # print(data)     #从excel中读取到的数据除了url之外，其它数据只有数字和字符串两种格式，所以经常需要转换数据格式以便后面调用
        data = eval(data)     #利用eval()函数对数据格式进行转换——字符串格式转换成字典格式
        # print(type(data))
        expected_result = test_case['expected_result']  #获取每一条测试用例的expected_result（期望结果）
        expected_result = expected_result.replace('null','None')  #将字符串中不可识别的内容进行替换
        expected_result = eval(expected_result)     #利用eval()函数对数据格式进行转换——字符串格式转换成字典格式
        # print(expected_result)
        # print(type(expected_result))
        real_result = post_func(qcd_url=url,qcd_data=data)    #调用发送接口请求的函数   请求参数data的数据格式必须是字典
        real_msg = real_result.get('msg')       #字典取值，获取要断言的有效字段
        # print(real_msg)
        expected_msg = expected_result.get('msg')
        print('真实执行结果是：{}'.format(real_msg))
        print('期望测试结果是：{}'.format(expected_msg))
        if real_msg == expected_msg:
            print('第{}条测试用例测试通过！'.format(case_id))
            final_result = 'Passed'    #给测试结果赋值一个变量，方便后期写入测试结果
        else:
            print('第{}条测试用例测试不通过！'.format(case_id))
            final_result = 'Failed'  # 给测试结果赋值一个变量，方便后期写入测试结果
        print('**'*20)
        wite_result(filename,sheetname,case_id+1,8,final_result)    #调用结果写入函数，进行测试结果写入

execute_func('test_case.xlsx','login')
execute_func('test_case.xlsx','register')
execute_func('test_case.xlsx','recharge')
























